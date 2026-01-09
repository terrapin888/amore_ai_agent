"""Excel 리포트 생성 모듈.

랭킹 데이터를 분석하여 Excel 형식의 리포트를 생성해요.
요약, 카테고리별 상세, LANEIGE 분석 시트를 포함해요.
"""

from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# 스타일 상수
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
LANEIGE_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
TOP5_FILL = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
BORDER = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))


class ExcelReportGenerator:
    """Excel 랭킹 리포트 생성기.

    랭킹 데이터를 분석하여 다양한 시트를 포함한
    Excel 리포트 파일을 생성해요.

    Attributes:
        output_dir (Path): 출력 디렉토리 경로
        workbook (Workbook | None): 현재 작업 중인 워크북
    """

    def __init__(self, output_dir: str = "output"):
        """ExcelReportGenerator를 초기화해요.

        Args:
            output_dir: 출력 디렉토리 경로 (기본값: "output")
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        self.workbook: Workbook | None = None

    def create_ranking_report(self, ranking_data: dict[str, pd.DataFrame], filename: str | None = None) -> str:
        """랭킹 리포트를 생성해요.

        Args:
            ranking_data: 카테고리별 랭킹 데이터프레임 딕셔너리
            filename: 출력 파일명 (기본값: None, 타임스탬프 기반 자동 생성)

        Returns:
            str: 생성된 파일 경로
        """
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"ranking_report_{timestamp}.xlsx"

        self.workbook = Workbook()
        if self.workbook.active:
            self.workbook.remove(self.workbook.active)

        self._create_summary_sheet(ranking_data)

        for category, df in ranking_data.items():
            if len(df) > 0:
                self._create_category_sheet(category, df)

        self._create_laneige_sheet(ranking_data)

        filepath = self.output_dir / filename
        assert self.workbook is not None
        self.workbook.save(filepath)

        return str(filepath)

    def _get_day_columns(self, df: pd.DataFrame) -> list:
        """데이터프레임에서 일별 컬럼을 정렬하여 반환해요.

        Args:
            df: 랭킹 데이터프레임

        Returns:
            list: 정렬된 day_N 컬럼 리스트
        """
        return sorted([c for c in df.columns if c.startswith("day_")], key=lambda x: int(x.split("_")[1]))

    def _create_summary_sheet(self, ranking_data: dict[str, pd.DataFrame]) -> None:
        """요약 시트를 생성해요.

        Args:
            ranking_data: 카테고리별 랭킹 데이터프레임 딕셔너리
        """
        assert self.workbook is not None
        ws = self.workbook.create_sheet("Summary", 0)

        ws["A1"] = "LANEIGE Ranking Report"
        ws["A1"].font = Font(size=16, bold=True)
        ws.merge_cells("A1:E1")

        ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws["A2"].font = Font(size=10, italic=True)

        row = 4
        ws[f"A{row}"] = "Category Summary"
        ws[f"A{row}"].font = Font(size=12, bold=True)
        row += 1

        headers = ["Category", "Total Products", "LANEIGE Products", "LANEIGE Best Rank", "LANEIGE Avg Rank"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

        row += 1
        for category, df in ranking_data.items():
            if len(df) == 0:
                continue

            day_cols = self._get_day_columns(df)
            laneige_df = df[df["is_laneige"]]
            total_products = len(df)
            laneige_products = len(laneige_df)

            best_rank: int | str = "-"
            avg_rank: float | str = "-"
            if len(laneige_df) > 0 and day_cols:
                all_ranks: list[float] = []
                for _, row_data in laneige_df.iterrows():
                    for col in day_cols:
                        if pd.notna(row_data[col]):
                            all_ranks.append(float(row_data[col]))
                if all_ranks:
                    best_rank = int(min(all_ranks))
                    avg_rank = round(sum(all_ranks) / len(all_ranks), 1)

            ws.cell(row=row, column=1, value=category.replace("_", " ").title())
            ws.cell(row=row, column=2, value=total_products)
            ws.cell(row=row, column=3, value=laneige_products)
            ws.cell(row=row, column=4, value=best_rank)
            ws.cell(row=row, column=5, value=avg_rank)

            if laneige_products > 0:
                for col in range(1, 6):
                    ws.cell(row=row, column=col).fill = LANEIGE_FILL

            row += 1

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 18

    def _create_category_sheet(self, category: str, df: pd.DataFrame) -> None:
        """카테고리별 상세 시트를 생성해요.

        Args:
            category: 카테고리명
            df: 해당 카테고리의 랭킹 데이터프레임
        """
        assert self.workbook is not None
        sheet_name = category.replace("_", " ").title()[:31]
        ws = self.workbook.create_sheet(sheet_name)

        day_cols = self._get_day_columns(df)

        today = datetime.now()
        date_labels = []
        for _i, col in enumerate(day_cols):
            day_num = int(col.split("_")[1])
            date = today - timedelta(days=len(day_cols) - day_num)
            date_labels.append(date.strftime("%m/%d"))

        headers = ["Product", "Brand", "LANEIGE"] + date_labels
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = BORDER

        for row_idx, (_, row_data) in enumerate(df.iterrows(), 2):
            is_laneige = row_data.get("is_laneige", False)

            ws.cell(row=row_idx, column=1, value=row_data.get("product_name", "")[:50])
            ws.cell(row=row_idx, column=2, value=row_data.get("brand", "")[:20])
            ws.cell(row=row_idx, column=3, value="Yes" if is_laneige else "No")

            for col_idx, day_col in enumerate(day_cols, 4):
                value = row_data.get(day_col)
                cell = ws.cell(row=row_idx, column=col_idx)

                if pd.notna(value):
                    cell.value = int(value)
                    if value <= 5:
                        cell.font = Font(bold=True, color="006400")
                    elif value <= 10:
                        cell.font = Font(bold=True, color="0000FF")
                else:
                    cell.value = "-"

                cell.border = BORDER
                cell.alignment = Alignment(horizontal="center")

            if is_laneige:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col).fill = LANEIGE_FILL

            for col in range(1, 4):
                ws.cell(row=row_idx, column=col).border = BORDER

        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 10

        if len(day_cols) > 0:
            for col_idx in range(4, 4 + len(day_cols)):
                col_letter = ws.cell(row=1, column=col_idx).column_letter
                range_str = f"{col_letter}2:{col_letter}{len(df) + 1}"

                color_scale = ColorScaleRule(
                    start_type="num",
                    start_value=1,
                    start_color="63BE7B",
                    mid_type="num",
                    mid_value=25,
                    mid_color="FFEB84",
                    end_type="num",
                    end_value=50,
                    end_color="F8696B",
                )
                ws.conditional_formatting.add(range_str, color_scale)

    def _create_laneige_sheet(self, ranking_data: dict[str, pd.DataFrame]) -> None:
        """LANEIGE 제품 분석 시트를 생성해요.

        Args:
            ranking_data: 카테고리별 랭킹 데이터프레임 딕셔너리
        """
        assert self.workbook is not None
        ws = self.workbook.create_sheet("LANEIGE Analysis")

        ws["A1"] = "LANEIGE Products Performance"
        ws["A1"].font = Font(size=14, bold=True)
        ws.merge_cells("A1:F1")

        row = 3

        for category, df in ranking_data.items():
            if len(df) == 0:
                continue

            laneige_df = df[df["is_laneige"]]
            if len(laneige_df) == 0:
                continue

            day_cols = self._get_day_columns(df)

            ws[f"A{row}"] = f"[{category.replace('_', ' ').title()}]"
            ws[f"A{row}"].font = Font(size=12, bold=True)
            row += 1

            headers = ["Product", "Best Rank", "Worst Rank", "Avg Rank", "TOP 5 Days", "TOP 10 Days"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.border = BORDER

            row += 1

            for _, product_row in laneige_df.iterrows():
                product_name = product_row.get("product_name", "Unknown")

                ranks = []
                for col in day_cols:
                    if pd.notna(product_row[col]):
                        ranks.append(product_row[col])

                if not ranks:
                    continue

                best_rank = int(min(ranks))
                worst_rank = int(max(ranks))
                avg_rank = round(sum(ranks) / len(ranks), 1)
                top5_days = sum(1 for r in ranks if r <= 5)
                top10_days = sum(1 for r in ranks if r <= 10)

                ws.cell(row=row, column=1, value=product_name[:50]).border = BORDER
                ws.cell(row=row, column=2, value=best_rank).border = BORDER
                ws.cell(row=row, column=3, value=worst_rank).border = BORDER
                ws.cell(row=row, column=4, value=avg_rank).border = BORDER
                ws.cell(row=row, column=5, value=top5_days).border = BORDER
                ws.cell(row=row, column=6, value=top10_days).border = BORDER

                if top5_days >= 7:
                    for col in range(1, 7):
                        ws.cell(row=row, column=col).fill = LANEIGE_FILL

                row += 1

            row += 2

        ws.column_dimensions["A"].width = 40
        for col_letter in "BCDEF":
            ws.column_dimensions[col_letter].width = 15


def generate_report(ranking_data: dict[str, pd.DataFrame], output_dir: str = "output") -> str:
    """랭킹 리포트를 생성하는 헬퍼 함수예요.

    Args:
        ranking_data: 카테고리별 랭킹 데이터프레임 딕셔너리
        output_dir: 출력 디렉토리 경로 (기본값: "output")

    Returns:
        str: 생성된 파일 경로
    """
    generator = ExcelReportGenerator(output_dir)
    return generator.create_ranking_report(ranking_data)
